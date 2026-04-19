import csv
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

# Paths
csv_path = 'lbc-rates.csv'
xlsx_path = 'lbc-rates.xlsx'

# Check if CSV exists
if not os.path.exists(csv_path):
    print(f"Error: {csv_path} not found")
    exit(1)

print(f"Reading {csv_path}...")

# Read CSV data
with open(csv_path, 'r', encoding='utf-8') as f:
    reader = csv.DictReader(f)
    data = list(reader)

print(f"✅ Read {len(data)} records")

# Create workbook
print(f"Creating {xlsx_path}...")
wb = Workbook()

# Sheet 1: All Data
ws_all = wb.active
ws_all.title = "All Rates"

# Write headers
headers = list(data[0].keys())
for col_num, header in enumerate(headers, 1):
    cell = ws_all.cell(row=1, column=col_num, value=header)
    cell.font = cell.font.copy(bold=True)

# Write data
for row_num, row_data in enumerate(data, 2):
    for col_num, header in enumerate(headers, 1):
        ws_all.cell(row=row_num, column=col_num, value=row_data.get(header, ''))

# Auto-adjust column widths
for column in ws_all.columns:
    max_length = 0
    column = [cell for cell in column]
    for cell in column:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    adjusted_width = (max_length + 2)
    if adjusted_width > 50:
        adjusted_width = 50
    ws_all.column_dimensions[column[0].column_letter].width = adjusted_width

# Save workbook
wb.save(xlsx_path)
print(f"✅ Successfully created {xlsx_path}")
print(f"📊 Total records: {len(data)}")
print(f"📁 File saved to: {os.path.abspath(xlsx_path)}")